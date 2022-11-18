from PIL import Image as PImage
from PIL.ExifTags import TAGS
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border,Side,Alignment,Font
import os

class Sheet():
    wb=Workbook()
    cursor:dict={}
    page:dict={}
    imgNo:dict={}
    doubleBorder=Side(border_style="double",color="000000")
    thinBorder=Side(border_style="thin",color="000000")
    thickBorder=Side(border_style="thick",color="000000")
    def insert_img(self,sheet_name,img_path):
        #rotate if needed
        Pimg=PImage.open(img_path)
        exifinfo = Pimg._getexif()
        ret={}
        if exifinfo != None:
            for tag, value in exifinfo.items():
                decoded = TAGS.get(tag, tag)
                ret[decoded] = value
        if "Orientation" in ret:
            if ret["Orientation"] == 3:
                Pimg = Pimg.rotate(180, expand=True)
                Pimg.save("./"+str(Path(img_path).parent)+"/"+str(Path(img_path).stem)+"_R"+".jpg")
                img_path=str(Path(img_path).parent)+"/"+str(Path(img_path).stem)+"_R"+".jpg"
            elif ret["Orientation"] == 6:
                Pimg = Pimg.rotate(270, expand=True)
                Pimg.save("./"+str(Path(img_path).parent)+"/"+str(Path(img_path).stem)+"_R"+".jpg")
                img_path=str(Path(img_path).parent)+"/"+str(Path(img_path).stem)+"_R"+".jpg"
            elif ret["Orientation"] == 8:
                Pimg = Pimg.rotate(90, expand=True)
                Pimg.save("./"+str(Path(img_path).parent)+"/"+str(Path(img_path).stem)+"_R"+".jpg")
                img_path=str(Path(img_path).parent)+"/"+str(Path(img_path).stem)+"_R"+".jpg"
        #img insert
        img=Image(img=img_path)
        self.img_resize(img)
        self.wb[sheet_name].add_image(img,self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]).coordinate)
    def add_sheet(self,sheet_name):
        self.wb.create_sheet(sheet_name)
        self.wb[sheet_name].title=sheet_name
        self.cursor[sheet_name]=[1,2]
        self.page[sheet_name]=0
        self.imgNo[sheet_name]=0
        return
    def add_title_section(self,sheet_name,img_path,total_page):
        if sheet_name not in self.wb.sheetnames:
            return
        self.page[sheet_name]+=1
        self.cursor[sheet_name][0]+=1
        #logo
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0],
                                        end_row=self.cursor[sheet_name][0]+2,
                                        start_column=self.cursor[sheet_name][1],
                                        end_column=self.cursor[sheet_name][1]+1,
                                        )
        #add description text
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+2).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+2).coordinate].font=Font(size = 16,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+2).coordinate].value="施工紀錄照片\nPhoto Record of Construction"

        #description
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0],
                                        end_row=self.cursor[sheet_name][0]+2,
                                        start_column=self.cursor[sheet_name][1]+2,
                                        end_column=self.cursor[sheet_name][1]+6,
                                        )
        #add page text
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+7).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+7).coordinate].font=Font(size = 16,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+7).coordinate].value="Rev.A\nPage"+str(self.page[sheet_name])+" of "+str(total_page)
        #page
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0],
                                        end_row=self.cursor[sheet_name][0]+2,
                                        start_column=self.cursor[sheet_name][1]+7,
                                        end_column=self.cursor[sheet_name][1]+8,
                                        )
        #border
        for x in range(self.cursor[sheet_name][0],self.cursor[sheet_name][0]+3):
            for y in range(self.cursor[sheet_name][1],self.cursor[sheet_name][1]+9):
                self.wb[sheet_name].cell(row=x, column=y).border=Border(top=self.thinBorder,left=self.thinBorder,bottom=self.thinBorder,right=self.thinBorder)
        #add logo img
        img=Image(img_path)
        img.width,img.height=(60,60)
        self.wb[sheet_name].add_image(img,self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]).coordinate)
        self.cursor[sheet_name][0]+=4
        return
    def merge_img_space(self,sheet_name):
        #img space
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0],
                                        end_row=self.cursor[sheet_name][0]+4,
                                        start_column=self.cursor[sheet_name][1],
                                        end_column=self.cursor[sheet_name][1]+5,
                                        )
        #photo NO. text(left)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+6).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+6).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+6).coordinate].value="照片編號:\nPhoto NO."
        #photo NO. text(right)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+7).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+7).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0], column=self.cursor[sheet_name][1]+7).coordinate].value=str(self.imgNo[sheet_name]).rjust(3,"0")
        #photo NO.
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0],
                                        end_row=self.cursor[sheet_name][0],
                                        start_column=self.cursor[sheet_name][1]+7,
                                        end_column=self.cursor[sheet_name][1]+8,
                                        )
    def merge_date_space(self,sheet_name,date):
        #Date text(left)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+1, column=self.cursor[sheet_name][1]+6).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+1, column=self.cursor[sheet_name][1]+6).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+1, column=self.cursor[sheet_name][1]+6).coordinate].value="拍攝日期:\nDate"
        #Date text(right)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+1, column=self.cursor[sheet_name][1]+7).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+1, column=self.cursor[sheet_name][1]+7).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+1, column=self.cursor[sheet_name][1]+7).coordinate].value=str(date)
        #Date
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0]+1,
                                        end_row=self.cursor[sheet_name][0]+1,
                                        start_column=self.cursor[sheet_name][1]+7,
                                        end_column=self.cursor[sheet_name][1]+8,
                                        )
    def merge_loc_space(self,sheet_name,location):
        #Location text(left)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+2, column=self.cursor[sheet_name][1]+6).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+2, column=self.cursor[sheet_name][1]+6).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+2, column=self.cursor[sheet_name][1]+6).coordinate].value="拍攝地點:\nLocation"
        #Location text(right)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+2, column=self.cursor[sheet_name][1]+7).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+2, column=self.cursor[sheet_name][1]+7).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+2, column=self.cursor[sheet_name][1]+7).coordinate].value=str(location)
        #Location
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0]+2,
                                        end_row=self.cursor[sheet_name][0]+2,
                                        start_column=self.cursor[sheet_name][1]+7,
                                        end_column=self.cursor[sheet_name][1]+8,
                                        )
    def merge_descript_space(self,sheet_name,description):
        #Description text(left)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+3, column=self.cursor[sheet_name][1]+6).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+3, column=self.cursor[sheet_name][1]+6).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+3, column=self.cursor[sheet_name][1]+6).coordinate].value="說明:\nDescription"
        #Description text(right)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+3, column=self.cursor[sheet_name][1]+7).coordinate].alignment=Alignment(horizontal="center",vertical="center",wrapText=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+3, column=self.cursor[sheet_name][1]+7).coordinate].font=Font(size = 12,bold=True)
        self.wb[sheet_name][self.wb[sheet_name].cell(row=self.cursor[sheet_name][0]+3, column=self.cursor[sheet_name][1]+7).coordinate].value=str(description)
        #Description
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0]+3,
                                        end_row=self.cursor[sheet_name][0]+4,
                                        start_column=self.cursor[sheet_name][1]+7,
                                        end_column=self.cursor[sheet_name][1]+8,
                                        )
        self.wb[sheet_name].merge_cells(start_row=self.cursor[sheet_name][0]+3,
                                        end_row=self.cursor[sheet_name][0]+4,
                                        start_column=self.cursor[sheet_name][1]+6,
                                        end_column=self.cursor[sheet_name][1]+6,
                                        )
    def img_resize(self,img):
        cell_size=[277,347]
        #ratio for width and height
        ratio=[img.height/cell_size[0],img.width/cell_size[1]]
        if ratio[0]>ratio[1]:
            img.height/=ratio[0]
            img.width/=ratio[0]
        else :
            img.height /= ratio[1]
            img.width /= ratio[1]
    def add_img_section(self,sheet_name,img_path,date,location,description):
        if sheet_name not in self.wb.sheetnames:
            return
        self.imgNo[sheet_name]+=1
        self.merge_img_space(sheet_name)
        self.merge_date_space(sheet_name,date)
        self.merge_loc_space(sheet_name,location)
        self.merge_descript_space(sheet_name,description)
        #border style
        for x in range(self.cursor[sheet_name][0],self.cursor[sheet_name][0]+5):
            for y in range(self.cursor[sheet_name][1],self.cursor[sheet_name][1]+9):
                self.wb[sheet_name].cell(row=x, column=y).border=Border(top=self.thinBorder,left=self.thinBorder,bottom=self.thinBorder,right=self.thinBorder)
        


        self.insert_img(sheet_name,img_path)   
        
        self.cursor[sheet_name][0]+=5
        return
    

if __name__ == '__main__':
    s=Sheet()
    s.add_sheet("cook")
    s.add_title_section("cook",img_path="img/logo.png",total_page=2)
    s.add_img_section("cook",img_path="img/gi.jpg",date="2022/10/21",location="草原",description="擷取自網路圖片")
    s.add_img_section("cook",img_path="img/bottle.jpg",date="2022/10/21",location="山下",description="擷取自網路圖片")
    s.add_img_section("cook",img_path="img/cat1.jpg",date="2022/10/21",location="跑道旁",description="擷取自網路圖片")
    s.add_title_section("cook",img_path="img/logo.png",total_page=2)
    s.add_img_section("cook",img_path="img/dog1.jpg",date="2022/10/21",location="草原",description="擷取自網路圖片")
    s.add_img_section("cook",img_path="img/dog2.jpg",date="2022/10/21",location="草原",description="擷取自網路圖片")
    s.add_img_section("cook",img_path="img/dog3.jpg",date="2022/10/21",location="草原",description="擷取自網路圖片")
    s.add_sheet("cool")
    s.add_img_section("cool",img_path="img/mouseV.jpg",date="2022/10/21",location="草原",description="擷取自網路圖片")
    s.wb.save("better.xlsx")